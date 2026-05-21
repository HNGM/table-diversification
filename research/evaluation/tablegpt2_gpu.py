"""
tablegpt2_gpu.py
================
Self-contained, GPU-accelerated TableGPT2 runner (HuggingFace Transformers backend).

Just copy this single file onto another machine and run it. It depends only on
third-party PyPI packages — nothing from the original `table-diversification`
repo is imported, and **no Docker is required**.

Model
-----
Defaults to `tablegpt/TableGPT2-7B` (Qwen2-based, ChatML format). The official
Qwen tokenizer uses `apply_chat_template(...)` so we let it render the prompt
instead of hand-rolling tags like the TableLLM/Llama-2 variant.

Target hardware
---------------
Tested for: Tesla T4 (15 GB), CUDA 12.2, driver 535.x, compute capability 7.5.

Install
-------
    pip install --index-url https://download.pytorch.org/whl/cu121 "torch>=2.5,<2.7"
    pip install transformers accelerate bitsandbytes \
        huggingface_hub tqdm pandas openpyxl numpy sentencepiece

The model is loaded in **4-bit (NF4)** via bitsandbytes so the 7B base
weights fit in ~5 GB on the T4, leaving plenty of room for the KV cache.

Sandbox
-------
The generated python code is executed in a **subprocess** (no Docker).
Isolation is best-effort:
  * a fresh empty working directory per query (under --work-dir)
  * a wall-clock timeout (--sandbox-timeout)
  * on Linux/macOS the child is put in its own process group so the timeout
    kills the whole tree; on POSIX we also apply soft resource limits
    (CPU time and address-space) via the `resource` module.
This is suitable for benign LLM-generated table-analysis code, NOT for
running untrusted/adversarial code.

Usage
-----
    python tablegpt2_gpu.py \
        --input-file  /path/to/dataset.json \
        --output-file /path/to/output.json \
        --data-root   /path/to/dataset_root   # 'data_file' paths resolved against this

Example
-------
    python tablegpt2_gpu.py --input-file dataset/original.json \
        --output-file results/tablegpt2_original.json --data-root ./ --resume
"""

from __future__ import annotations

import argparse
import atexit
import datetime
import json
import os
import re
import shutil
import signal
import subprocess
import sys
import tempfile
import time
import traceback
from hashlib import sha256
from pathlib import Path
from typing import List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

import torch
from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
# Official TableGPT2 release (Qwen2-7B fine-tune).
GPU_MODEL_REPO_ID = "tablegpt/TableGPT2-7B"
GPU_MAX_NEW_TOKENS = 1024
GPU_MAIN_DEVICE = "cuda:0"
GPU_LOAD_IN_4BIT = True   # set False to load in fp16 (~13 GB on the T4)

PROMPT_MODE = "default"  # set to "mistake" to include the disturbance hint

SANDBOX_TIMEOUT_SEC = 150
SANDBOX_MEM_LIMIT_MB = 4096       # POSIX only (RLIMIT_AS); 0 disables
SANDBOX_CPU_LIMIT_SEC = 180       # POSIX only (RLIMIT_CPU); 0 disables

MISTAKE_PROMPT = (
    "If you encounter scnearios where the table seems incorrect either "
    "structurally or semantically (e.g., shifted rows, shifted columns, "
    "semantic misalignment, etc.), ensure that your code accounts for "
    "these inconsistencies to provide an accurate and robust answer."
)

# TableGPT2 uses ChatML (system + user roles). Code-generation prompt is sent
# as the user turn. The placeholder {data_file} is filled with the actual
# uploaded filename, so the model is told exactly which file to read.
TABLEGPT2_PROMPT = """Below is the table data. You need to write a Python program that reads the data as 'data.xlsx' to solve the problem.

{mistake_instr}

Generate a python script to solve the above in the following format:
```python
<python code>
```

DATA CONTEXT:
{data_preview}

Question: {user_question}
"""


# ---------------------------------------------------------------------------
# Transformers backend: a tiny adapter exposing `create_chat_completion`
# so the rest of the file is unchanged.
# ---------------------------------------------------------------------------
class HFChatLLM:
    """Thin wrapper around AutoModelForCausalLM that mimics the
    `llama_cpp.Llama.create_chat_completion(messages=...)` interface.

    Uses the tokenizer's built-in chat template (TableGPT2 / Qwen2 ChatML).
    """

    def __init__(
        self,
        model_id: str,
        device: str = GPU_MAIN_DEVICE,
        load_in_4bit: bool = GPU_LOAD_IN_4BIT,
        max_new_tokens: int = GPU_MAX_NEW_TOKENS,
    ):
        if not torch.cuda.is_available():
            raise RuntimeError(
                "CUDA is not available to PyTorch. Verify driver + cu121 torch wheel."
            )
        print(f"[HFChatLLM] CUDA device: {torch.cuda.get_device_name(0)}")

        self.max_new_tokens = max_new_tokens
        self.device = device

        self.tokenizer = AutoTokenizer.from_pretrained(model_id, use_fast=True)
        if self.tokenizer.pad_token_id is None:
            self.tokenizer.pad_token_id = self.tokenizer.eos_token_id

        kwargs = dict(device_map={"": device}, torch_dtype=torch.float16)
        if load_in_4bit:
            kwargs["quantization_config"] = BitsAndBytesConfig(
                load_in_4bit=True,
                bnb_4bit_quant_type="nf4",
                bnb_4bit_compute_dtype=torch.float16,
                bnb_4bit_use_double_quant=True,
            )
            # device_map must be auto/single device — let accelerate place it.
            kwargs["device_map"] = {"": 0}
            kwargs.pop("torch_dtype", None)

        self.model = AutoModelForCausalLM.from_pretrained(
            model_id, trust_remote_code=True, **kwargs
        )
        self.model.eval()

    def _messages_to_prompt(self, messages: List[dict]) -> str:
        """Render via the tokenizer's chat template (ChatML for TableGPT2)."""
        return self.tokenizer.apply_chat_template(
            messages, tokenize=False, add_generation_prompt=True
        )

    @torch.inference_mode()
    def create_chat_completion(self, messages: List[dict]) -> dict:
        prompt = self._messages_to_prompt(messages)
        inputs = self.tokenizer(prompt, return_tensors="pt").to(self.device)
        out = self.model.generate(
            **inputs,
            max_new_tokens=self.max_new_tokens,
            do_sample=False,
            temperature=1.0,
            pad_token_id=self.tokenizer.pad_token_id,
            eos_token_id=self.tokenizer.eos_token_id,
        )
        new_tokens = out[0, inputs["input_ids"].shape[1]:]
        text = self.tokenizer.decode(new_tokens, skip_special_tokens=True)
        return {"choices": [{"message": {"role": "assistant", "content": text}}]}


# ---------------------------------------------------------------------------
# JSON I/O
# ---------------------------------------------------------------------------
def read_json(filepath: Union[Path, str]):
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        return json.load(f)


def write_json(data, filepath: Union[Path, str]):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


# ---------------------------------------------------------------------------
# Data preview (markdown) — supports .csv / .xls / .xlsx
# ---------------------------------------------------------------------------
def _worksheet_to_dataframe(ws) -> pd.DataFrame:
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell) for cell in row])
    return pd.DataFrame(rows)


def get_data_preview_markdown(data_file: Path) -> str:
    data_file = Path(data_file)
    suffix = data_file.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(data_file, dtype=str).fillna("")
        headers = list(df.columns)
        body_rows = [
            "| " + " | ".join(str(v) for v in row.tolist()) + " |"
            for _, row in df.iterrows()
        ]
    elif suffix in {".xls", ".xlsx"}:
        wb = load_workbook(data_file, data_only=True)
        ws = wb.active
        df = _worksheet_to_dataframe(ws)
        headers = df.iloc[0].tolist()
        df = df.iloc[1:].reset_index(drop=True)
        headers = ["" if str(h).startswith("Unnamed") else h for h in headers]
        body_rows = [
            "| " + " | ".join(row.tolist()) + " |"
            for _, row in df.iterrows()
        ]
    else:
        return "No preview available for this data.\n"

    header_row = "| " + " | ".join(str(h) for h in headers) + " |"
    separator_row = "| " + " | ".join(["---"] * len(headers)) + " |"
    return "\n".join([header_row, separator_row] + body_rows)


# ---------------------------------------------------------------------------
# Subprocess-based code sandbox (no Docker)
# ---------------------------------------------------------------------------
class CodeToolRequest:
    """Extract the python code from a possibly fenced model response."""

    def __init__(self, code_str: str):
        code_blocks = re.findall(
            r"```(?:python)?(?:py)?\s*(.*?)\s*```", code_str, re.DOTALL
        )
        self.code = "\n".join(code_blocks).strip() if code_blocks else code_str.strip()

        # Ensure the last expression is printed if it isn't already.
        lines = self.code.splitlines(keepends=True)
        if lines and "print" not in lines[-1]:
            last_line = lines[-1]
            last_line_lstrip = last_line.lstrip()
            indent = last_line[: last_line.index(last_line_lstrip)]
            line_content = last_line_lstrip.strip()
            if line_content and not line_content.startswith("#"):
                tail = last_line.split(line_content, 2)[1]
                lines[-1] = indent + f"print({line_content})" + tail
                self.code = "".join(lines)


class CodeToolResponse:
    def __init__(self, exit_code: int, log: str, output_files: List[Path]):
        self.exit_code = exit_code
        self.log = log
        self.output_files = output_files


def _posix_preexec(mem_mb: int, cpu_sec: int):
    """preexec_fn for POSIX: new process group + soft resource limits."""
    import resource  # POSIX only

    os.setsid()
    if cpu_sec and cpu_sec > 0:
        resource.setrlimit(resource.RLIMIT_CPU, (cpu_sec, cpu_sec))
    if mem_mb and mem_mb > 0:
        bytes_ = mem_mb * 1024 * 1024
        try:
            resource.setrlimit(resource.RLIMIT_AS, (bytes_, bytes_))
        except (ValueError, OSError):
            pass  # macOS rejects RLIMIT_AS in some cases


class CodeTool:
    """Run python code in an isolated subprocess.

    Each instance owns a fresh working directory. Uploaded data files are
    copied into it, then `python <script>` is invoked there.
    """

    def __init__(
        self,
        time_out: int = SANDBOX_TIMEOUT_SEC,
        mem_limit_mb: int = SANDBOX_MEM_LIMIT_MB,
        cpu_limit_sec: int = SANDBOX_CPU_LIMIT_SEC,
        work_root: Union[str, Path] = None,
        python_executable: Optional[str] = None,
    ):
        self._time_out = time_out
        self._mem_limit_mb = mem_limit_mb
        self._cpu_limit_sec = cpu_limit_sec
        self._python = python_executable or sys.executable

        sandbox_id = sha256(str(time.time()).encode()).hexdigest()[:16]
        root = Path(work_root) if work_root else Path(tempfile.gettempdir()) / "tablegpt2_sandbox"
        self._work_dir = root / sandbox_id
        self._work_dir.mkdir(parents=True, exist_ok=True)

        self._uploaded_files: List[Path] = []
        self._generated_files: List[Path] = []

    def upload_files(self, file_paths: List[Path]) -> List[Path]:
        copied = []
        for fp in file_paths:
            fp = Path(fp)
            if not fp.exists():
                print(f"File {fp} does not exist")
                continue
            dest = self._work_dir / fp.name
            shutil.copyfile(fp, dest)
            copied.append(dest)
            self._uploaded_files.append(dest)
        return copied

    def run_code(self, code_str: str) -> CodeToolResponse:
        req = CodeToolRequest(code_str)
        if not req.code:
            return CodeToolResponse(1, "No code to execute", [])

        script_name = f"exec_{sha256(str(time.time()).encode()).hexdigest()[:12]}.py"
        script_path = self._work_dir / script_name
        script_path.write_text(req.code, encoding="utf-8")

        existing = set(self._work_dir.glob("*"))

        kwargs = dict(
            cwd=str(self._work_dir),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="ignore",
        )
        if os.name == "posix":
            kwargs["preexec_fn"] = lambda: _posix_preexec(
                self._mem_limit_mb, self._cpu_limit_sec
            )
        else:
            # Windows: new process group so we can kill the whole tree.
            kwargs["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP  # type: ignore[attr-defined]

        proc = subprocess.Popen([self._python, script_name], **kwargs)
        try:
            log, _ = proc.communicate(timeout=self._time_out)
            exit_code = proc.returncode
        except subprocess.TimeoutExpired:
            try:
                if os.name == "posix":
                    os.killpg(proc.pid, signal.SIGKILL)
                else:
                    proc.send_signal(signal.CTRL_BREAK_EVENT)  # type: ignore[attr-defined]
                    proc.kill()
            except Exception:
                pass
            try:
                log, _ = proc.communicate(timeout=5)
            except Exception:
                log = ""
            return CodeToolResponse(124, (log or "") + "\n[TIMEOUT]", [])

        new_files = [
            f for f in self._work_dir.glob("*")
            if f not in existing and f != script_path
        ]
        self._generated_files.extend(new_files)
        return CodeToolResponse(exit_code, (log or "").rstrip(), new_files)

    def cleanup(self):
        try:
            shutil.rmtree(self._work_dir, ignore_errors=True)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Core workflow: prompt -> code -> sandbox -> log
# ---------------------------------------------------------------------------
def _resolve_data_path(data_root: Path, raw_path: str) -> Path:
    p = Path(raw_path)
    return p if p.is_absolute() else (data_root / p)


def _rewrite_data_filename(code_block: str, data_file_name: str) -> str:
    """Replace the placeholder 'data.xlsx' with the real uploaded file name."""
    if code_block.count("'data.xlsx'") == 1:
        return code_block.replace("'data.xlsx'", f"'{data_file_name}'")
    return re.sub(
        r"pd\.read_excel\(\s*'data\.xlsx'\s*\)",
        f"pd.read_excel('{data_file_name}')",
        code_block,
    )


MAX_TRIES = 3


def tooling_workflow(
    llm: "HFChatLLM",
    info: dict,
    data_root: Path,
    sandbox_kwargs: dict,
    max_tries: int = MAX_TRIES,
) -> dict:
    """Prompt the model, run its python, retry on failure with error feedback.

    Up to `max_tries` model calls. On a non-zero sandbox exit code the error
    log is appended as a user message and the model is asked to fix the code.
    The final result records every attempt.
    """
    data_path = _resolve_data_path(data_root, info["data_file"])
    data_file_name = data_path.name
    markdown = get_data_preview_markdown(data_path)

    code_tool = CodeTool(**sandbox_kwargs)
    try:
        code_tool.upload_files([data_path])

        initial_prompt = (
            TABLEGPT2_PROMPT
            .replace("{data_preview}", markdown)
            .replace("{user_question}", info["query"])
            .replace("{mistake_instr}", MISTAKE_PROMPT if PROMPT_MODE == "mistake" else "")
        )
        messages = [{"role": "user", "content": initial_prompt}]

        attempts = []
        final_code_response: Optional[CodeToolResponse] = None
        final_code_block: Optional[str] = None
        final_raw_response: Optional[str] = None

        for attempt_idx in range(1, max_tries + 1):
            response = llm.create_chat_completion(messages=messages)
            raw_response = response["choices"][0]["message"]["content"]
            messages.append({"role": "assistant", "content": raw_response})

            python_matches = re.findall(
                r"```python\s*(.*?)\s*```", raw_response, re.DOTALL
            )
            code_block = python_matches[0] if python_matches else raw_response
            code_block = _rewrite_data_filename(code_block, data_file_name)

            code_response = code_tool.run_code(code_block)

            attempts.append({
                "attempt": attempt_idx,
                "raw_response": raw_response,
                "code": code_block,
                "code_response_log": code_response.log,
                "exit_code": code_response.exit_code,
            })

            final_code_response = code_response
            final_code_block = code_block
            final_raw_response = raw_response

            if code_response.exit_code == 0:
                messages.append({
                    "role": "user",
                    "content": f"Code block executed successfully:\n{code_response.log}",
                })
                break

            # Error path — feed the log back to the model for the next try.
            if attempt_idx < max_tries:
                messages.append({
                    "role": "user",
                    "content": (
                        f"Code block execution failed with exit code "
                        f"{code_response.exit_code}:\n{code_response.log}\n"
                        f"Please fix the code and try again."
                    ),
                })

        info["eval"] = [{
            "raw_response": final_raw_response,
            "code": final_code_block,
            "code_response_log": final_code_response.log if final_code_response else "",
            "exit_code": final_code_response.exit_code if final_code_response else -1,
            "attempts": attempts,
            "messages": messages,
            "tools": len(attempts),
        }]
        return info
    finally:
        code_tool.cleanup()


# ---------------------------------------------------------------------------
# Streaming checkpoint helpers
# ---------------------------------------------------------------------------
def _read_jsonl_lenient(jsonl_path: Path) -> list:
    results = []
    if not jsonl_path.exists():
        return results
    with open(jsonl_path, "r", encoding="utf-8") as f:
        for line_no, raw in enumerate(f, start=1):
            line = raw.strip()
            if not line:
                continue
            try:
                results.append(json.loads(line))
            except json.JSONDecodeError as e:
                print(f"Warning: skipping malformed line {line_no} in {jsonl_path}: {e}")
    return results


def _dedupe_by_index(results: list) -> list:
    by_index, order = {}, []
    for item in results:
        idx = item.get("index")
        if idx is None:
            order.append(id(item))
            by_index[id(item)] = item
            continue
        if idx not in by_index:
            order.append(idx)
        by_index[idx] = item
    return [by_index[k] for k in order]


def _finalize_jsonl_to_json(jsonl_path: Path, json_path: Path):
    if not jsonl_path.exists():
        return
    results = _read_jsonl_lenient(jsonl_path)
    if not results and json_path.exists():
        print(f"No results in {jsonl_path}; leaving existing {json_path} untouched.")
        try:
            jsonl_path.unlink()
        except OSError as e:
            print(f"Warning: failed to delete {jsonl_path}: {e}")
        return
    results = _dedupe_by_index(results)
    try:
        write_json(results, json_path)
        print(f"Finalized {len(results)} results to {json_path}")
    except Exception as e:
        print(f"Failed to finalize jsonl -> json: {e}")
        return
    try:
        jsonl_path.unlink()
    except OSError as e:
        print(f"Warning: failed to delete {jsonl_path}: {e}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def get_config(args):
    here = Path(__file__).resolve().parent
    default_out = here / "results" / datetime.datetime.now().strftime("%d%m%y") / "tablegpt2_gpu_output.json"

    parser = argparse.ArgumentParser(description="Standalone GPU TableGPT2 runner (no Docker)")
    parser.add_argument("--input-file", required=True,
                        help="JSON file containing dataset entries (each must have 'index', 'query', 'data_file').")
    parser.add_argument("--output-file", default=default_out)
    parser.add_argument("--data-root", default=None,
                        help="Root dir against which relative 'data_file' paths are resolved. "
                             "Defaults to the parent of --input-file.")
    parser.add_argument("--resume", action="store_true", help="Resume from the last checkpoint")

    # Model knobs
    parser.add_argument("--model-repo-id", default=GPU_MODEL_REPO_ID,
                        help="HuggingFace repo id for the TableGPT2 model.")
    parser.add_argument("--max-new-tokens", type=int, default=GPU_MAX_NEW_TOKENS,
                        help="Maximum new tokens generated per model call.")
    parser.add_argument("--device", default=GPU_MAIN_DEVICE,
                        help="CUDA device for model placement, e.g. 'cuda:0'.")
    parser.add_argument("--no-4bit", action="store_true",
                        help="Load the model in fp16 instead of 4-bit (needs ~13 GB VRAM).")

    # Sandbox knobs
    parser.add_argument("--sandbox-timeout", type=int, default=SANDBOX_TIMEOUT_SEC,
                        help="Wall-clock timeout (seconds) per generated script.")
    parser.add_argument("--sandbox-mem-mb", type=int, default=SANDBOX_MEM_LIMIT_MB,
                        help="POSIX-only memory cap (MB) via RLIMIT_AS. 0 disables.")
    parser.add_argument("--sandbox-cpu-sec", type=int, default=SANDBOX_CPU_LIMIT_SEC,
                        help="POSIX-only CPU-time cap (seconds) via RLIMIT_CPU. 0 disables.")
    parser.add_argument("--sandbox-python", default=None,
                        help="Python executable used in the sandbox. Defaults to the current interpreter.")
    parser.add_argument("--work-dir", default=None,
                        help="Root directory under which per-query sandbox dirs are created. "
                             "Defaults to <tempdir>/tablegpt2_sandbox.")
    parser.add_argument("--max-tries", type=int, default=MAX_TRIES,
                        help="Maximum model attempts per query (error log is fed back between tries).")

    cfg = parser.parse_args(args)
    cfg.input_file = Path(cfg.input_file)
    cfg.output_file = Path(cfg.output_file)
    cfg.data_root = Path(cfg.data_root) if cfg.data_root else cfg.input_file.parent
    cfg.output_file.parent.mkdir(parents=True, exist_ok=True)
    return cfg


def main(args):
    cfg = get_config(args)

    llm = HFChatLLM(
        model_id=cfg.model_repo_id,
        device=cfg.device,
        load_in_4bit=not cfg.no_4bit,
        max_new_tokens=cfg.max_new_tokens,
    )

    sandbox_kwargs = dict(
        time_out=cfg.sandbox_timeout,
        mem_limit_mb=cfg.sandbox_mem_mb,
        cpu_limit_sec=cfg.sandbox_cpu_sec,
        work_root=cfg.work_dir,
        python_executable=cfg.sandbox_python,
    )

    dataset = read_json(cfg.input_file)

    jsonl_path = cfg.output_file.with_suffix(".jsonl")
    existing_results: list = []
    processed_indices: set = set()

    if cfg.resume:
        if jsonl_path.exists():
            print(f"Found leftover streaming checkpoint {jsonl_path}. Recovering it.")
            existing_results = _read_jsonl_lenient(jsonl_path)
            if cfg.output_file.exists():
                try:
                    finalized = read_json(cfg.output_file)
                    jsonl_indices = {it.get("index") for it in existing_results}
                    extra = [it for it in finalized if it.get("index") not in jsonl_indices]
                    if extra:
                        print(f"Merging {len(extra)} extra items from {cfg.output_file}.")
                        existing_results.extend(extra)
                except Exception as e:
                    print(f"Warning: could not merge {cfg.output_file}: {e}")
        elif cfg.output_file.exists():
            print(f"Output file {cfg.output_file} exists. Seeding jsonl checkpoint from it.")
            existing_results = read_json(cfg.output_file)
        else:
            print("Resume requested but no prior checkpoint or output file found. Starting fresh.")

        existing_results = _dedupe_by_index(existing_results)
        if existing_results:
            tmp_path = jsonl_path.with_suffix(jsonl_path.suffix + ".tmp")
            with open(tmp_path, "w", encoding="utf-8") as f:
                for item in existing_results:
                    f.write(json.dumps(item, ensure_ascii=False) + "\n")
            os.replace(tmp_path, jsonl_path)
        elif jsonl_path.exists():
            jsonl_path.unlink()

        processed_indices = {item["index"] for item in existing_results if "index" in item}
        print(f"Found {len(processed_indices)} already processed items.")
    else:
        if jsonl_path.exists():
            print(f"Fresh run: removing stale checkpoint {jsonl_path}.")
            jsonl_path.unlink()

    rem_dataset = [d for d in dataset if d["index"] not in processed_indices]
    print(f"Resuming evaluation. {len(rem_dataset)} out of {len(dataset)} remaining.")

    _finalized = {"done": False}

    def _finalize_once():
        if _finalized["done"]:
            return
        _finalized["done"] = True
        _finalize_jsonl_to_json(jsonl_path, cfg.output_file)

    atexit.register(_finalize_once)

    def _signal_handler(signum, frame):
        print(f"\nReceived signal {signum}. Finalizing results...")
        _finalize_once()
        sys.exit(128 + signum)

    for sig in (signal.SIGINT, signal.SIGTERM):
        try:
            signal.signal(sig, _signal_handler)
        except (ValueError, OSError):
            pass

    try:
        with open(jsonl_path, "a", encoding="utf-8") as ckpt_f:
            for data in tqdm(rem_dataset):
                try:
                    result = tooling_workflow(
                        llm, data, cfg.data_root, sandbox_kwargs,
                        max_tries=cfg.max_tries,
                    )
                except Exception as e:
                    print(f"Error processing index {data.get('index')}: {e}")
                    print(f"Traceback: {traceback.format_exc()}")
                    continue

                try:
                    line = json.dumps(result, ensure_ascii=False)
                except (TypeError, ValueError) as e:
                    print(f"Error serializing result for index {data.get('index')}: {e}")
                    continue

                ckpt_f.write(line + "\n")
                ckpt_f.flush()
    finally:
        _finalize_once()


if __name__ == "__main__":
    main(sys.argv[1:])
