"""Self-contained copy of `src.utils.data_preview.get_data_preview_markdown`.

Vendored here so the `research/finetuning/` folder can be copied/run on any
machine without depending on the rest of the repository. Logic is byte-identical
to the canonical helper used by the evaluation pipeline.
"""
from __future__ import annotations

from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook

_NO_PREVIEW_MESSAGE = "No preview available for this data.\n"


def _worksheet_to_dataframe(ws) -> pd.DataFrame:
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell) for cell in row])
    return pd.DataFrame(rows)


def get_data_preview_markdown(data_file: Path) -> str:
    data_file = Path(data_file)
    if data_file.suffix.lower() == ".csv":
        df = pd.read_csv(data_file, dtype=str).fillna("")
    elif data_file.suffix.lower() in {".xls", ".xlsx"}:
        wb = load_workbook(data_file, data_only=True)
        ws = wb.active
        df = _worksheet_to_dataframe(ws)
    else:
        return _NO_PREVIEW_MESSAGE

    # First row is treated as header row (even if broken)
    headers = df.iloc[0].tolist()
    df = df.iloc[1:].reset_index(drop=True)

    # Replace pandas-generated "Unnamed" headers or None
    headers = [
        "" if str(h).startswith("Unnamed") else h
        for h in headers
    ]

    # --- Manual markdown rendering ---
    header_row = "| " + " | ".join(headers) + " |"
    separator_row = "| " + " | ".join(["---"] * len(headers)) + " |"

    body_rows = []
    for _, row in df.iterrows():
        body_rows.append("| " + " | ".join(row.tolist()) + " |")

    return "\n".join([header_row, separator_row] + body_rows)
