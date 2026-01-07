from io import StringIO
from pathlib import Path
import warnings
import chardet
import numpy as np
import pandas as pd
from pandas.api.types import is_string_dtype, is_numeric_dtype
from scipy.stats import zscore
from tabulate import tabulate
from openpyxl import load_workbook
from typing import List


# root_dir = Path(__file__).parent.resolve()
# while not root_dir.name == "CoDABench":
#    root_dir = root_dir.parent
# PROJECT_DIR = root_dir
warnings.filterwarnings("ignore")

_NUM_SAMPLES = 20
_READ_FAILURE_MESSAGE = "Read Failure: {}, Reason: {}"
_PREVIEW_FAILURE_MESSAGE = "Preview Failure: {}, Reason: {}"
_NO_PREVIEW_MESSAGE = "No preview available for this data.\n"


def diversity_sample(df, num_samples:int):
    samples_per_column = num_samples // df.shape[1]
    sampled_df = pd.DataFrame(columns=df.columns)
    for col in df.columns:

        # if the column has any missing values, add min(1, .25 * samples_per_column) of those
        missing_val_rows = df[df[col].isnull() | (df[col] == "")]
        num_sampled_rows = 0
        if missing_val_rows.shape[0] > 0:
            df, samples = _sample_missing_values(
                df,
                missing_val_rows,
                samples_per_column
            )
            num_sampled_rows = samples.shape[0]

        df_missing_vals_removed = df[df[col].notnull()]
        if df_missing_vals_removed.shape[0] == 0:
            continue

        # If a column has a numeric type, we weight rows that are numerical outliers
        if is_numeric_dtype(df_missing_vals_removed[col]):
            df, samples = _sample_numeric_column(
                col,
                df,
                df_missing_vals_removed,
                samples_per_column,
                num_sampled_rows
            )
            sampled_df = pd.concat([sampled_df, samples], axis=0)

        # If the column has a string type, we weight rows that are outliers in terms of length AND/OR frequency
        elif is_string_dtype(df_missing_vals_removed[col]):
            df, samples = _sample_string_column(
                col,
                df,
                df_missing_vals_removed,
                samples_per_column,
                num_sampled_rows
            )
            sampled_df = pd.concat([sampled_df, samples])

        # If the column has an object type, we weight rows the are outliers in terms of frequency
        else:
            df, samples = _sample_generic_column(
                col,
                df,
                df_missing_vals_removed,
                samples_per_column,
                num_sampled_rows
            )
            sampled_df = pd.concat([sampled_df, samples])

    # If we do not have enough samples, randomly sample the remaining rows
    if sampled_df.shape[0] < num_samples and df.shape[0]:
        samples = df.sample(
            n=min(df.shape[0], max(num_samples - sampled_df.shape[0], 0)),
            random_state=1
        )
        sampled_df = pd.concat([sampled_df, samples])

    return sampled_df

def _sample_numeric_column(
    col,
    df,
    df_missing_vals_removed,
    samples_per_column,
    num_sampled_rows
):
    z_scores = zscore(df_missing_vals_removed[col])
    probs = np.abs(z_scores)
    probs /= probs.sum()

    probs = probs + .1
    probs = probs.fillna(.1)

    samples =  df_missing_vals_removed.sample(
        n=min(
            df_missing_vals_removed.shape[0],
            max(samples_per_column - num_sampled_rows, 0)
        ),
        weights=probs,
        random_state=1
    )

    # Merge the dataframes with indicator to find differences
    merged_df = df.merge(samples, on=list(df.columns), how='outer', indicator=True)

    # Filter rows that are only in df1
    df = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])

    return df, samples


def _sample_string_column(
    col,
    df,
    df_missing_vals_removed,
    samples_per_column,
    num_sampled_rows
):
    # Calculate the length of each string in col_y
    str_lengths = df_missing_vals_removed[col].apply(len).astype(int)

    # Calculate the frequency of each string in col_y
    frequency = df_missing_vals_removed[col].value_counts()
    str_freqs = df_missing_vals_removed[col].map(frequency).astype(int)
    # print('str freqs:')
    # print(str_freqs)
    # print(frequency)

    # Calculate z-scores for length and frequency
    length_z_scores = zscore(str_lengths)
    freq_z_scores = zscore(str_freqs)

    # Combine the z-scores to create a probability distribution
    # We use the sum of absolute z-scores to emphasize outliers
    combined_score = np.abs(length_z_scores) + np.abs(freq_z_scores)
    probs = combined_score / combined_score.sum()

    probs = probs + .1
    probs = probs.fillna(.1)

    # Sample rows using the probability distribution
    samples = df_missing_vals_removed.sample(
        n=min(
            df_missing_vals_removed.shape[0],
            max(samples_per_column - num_sampled_rows, 0)
        ),
        weights=probs,
        random_state=1
    )

    # Merge the dataframes with indicator to find differences
    merged_df = df.merge(samples, on=list(df.columns), how='outer', indicator=True)

    # Filter rows that are only in df1
    df = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])

    return df, samples

def _sample_generic_column(
    col,
    df,
    df_missing_vals_removed,
    samples_per_column,
    num_sampled_rows
):
    # Calculate the frequency of each string in col_y
    frequency = df_missing_vals_removed[col].value_counts()
    str_freqs = df_missing_vals_removed[col].map(frequency)

    # Calculate z-score for frequency
    z_scores = zscore(str_freqs)

    # Combine the z-scores to create a probability distribution
    # We use the sum of absolute z-scores to emphasize outliers
    probs = np.abs(z_scores)
    probs = probs / probs.sum()

    probs = probs + .1
    probs = probs.fillna(.1)

    # Sample rows using the probability distribution
    samples = df_missing_vals_removed.sample(
        n=min(
            df_missing_vals_removed.shape[0],
            max(samples_per_column - num_sampled_rows, 0)
        ),
        weights=probs,
        random_state=1
    )

    # Merge the dataframes with indicator to find differences
    merged_df = df.merge(samples, on=list(df.columns), how='outer', indicator=True)

    # Filter rows that are only in df1
    df = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])

    return df, samples


def _sample_missing_values(
    df,
    missing_val_rows,
    samples_per_column
):
    # if the column has any missing values, add min(1, .25 * samples_per_column) of those
    samples = missing_val_rows.sample(
        n = max(1, min(missing_val_rows.shape[0], samples_per_column // 4)),
        random_state=1
    )

    # Merge the dataframes with indicator to find differences
    merged_df = df.merge(samples, on=list(df.columns), how='outer', indicator=True)

    # Filter rows that are only in df1
    df = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])

    return samples, df



def get_data_preview(data_dir:Path, num_samples:int = _NUM_SAMPLES)->str:
    read_failure_count = 0
    preview_failure_count = 0
    data_dir = Path(data_dir)
    # check if data_dir is inside the root directory
    # if not data_dir.is_absolute():
    #    data_dir = Path(PROJECT_DIR) / data_dir
    try:
        if data_dir.suffix == ".csv":
            with open(data_dir, 'rb') as f:
                raw_data = f.read()
            # Detect encoding
            result = chardet.detect(raw_data)
            encoding = result['encoding'] if result['encoding'] else 'utf-8'
            # Decode the raw data
            text = raw_data.decode(encoding)
            # Read CSV from the decoded text
            sheet_df_dict = {data_dir.stem : pd.read_csv(StringIO(text))}
        else:
            sheet_df_dict = pd.read_excel(data_dir, sheet_name=None)
    except Exception as e:
        read_failure_count += 1
        print(_READ_FAILURE_MESSAGE.format(read_failure_count, e))
        return _NO_PREVIEW_MESSAGE
    
    output = ""        
    for sheet_name, df in sheet_df_dict.items():
        if len(df) <= num_samples:
            output += f"Sheet: {sheet_name}\n" + tabulate(df, headers="keys", tablefmt="grid") + "\n\n"
        else:
            try:
                sampled_df = diversity_sample(df, num_samples)
                output += f"Sheet: {sheet_name}\n" + tabulate(sampled_df, headers="keys", tablefmt="grid") + "\n\n"
            except Exception as e:
                output += f"Sheet: {sheet_name}\n" + _NO_PREVIEW_MESSAGE
                preview_failure_count += 1
                print(_PREVIEW_FAILURE_MESSAGE.format(preview_failure_count, e))

    return output

def _expand_merged_cells(ws):
    """
    Fill merged cell ranges so every cell contains the top-left value.
    This preserves visual structure for markdown rendering.
    """
    # Collect merged ranges first (to avoid modifying while iterating)
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        # Get the value from the top-left cell before unmerging
        value = ws.cell(row=min_row, column=min_col).value
        
        # Unmerge the cells
        ws.unmerge_cells(str(merged_range))
        
        # Now fill all cells with the value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).value = value

def _worksheet_to_dataframe(ws) -> pd.DataFrame:
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(
            ["" if cell is None else str(cell) for cell in row]
        )
    return pd.DataFrame(rows)

def get_data_preview_markdown(data_file: Path) -> str:
    if data_file.suffix.lower() == ".csv":
        df = pd.read_csv(data_file, dtype=str).fillna("")
    elif data_file.suffix.lower() in {".xls", ".xlsx"}:
        wb = load_workbook(data_file, data_only=True)
        ws = wb.active

        df = _worksheet_to_dataframe(ws)
    else:
        return _NO_PREVIEW_MESSAGE

    # First row is treated as header row (even if broken)
    headers = df.iloc[0].tolist()
    df = df.iloc[1:].reset_index(drop=True)

    # Replace pandas-generated "Unnamed" headers or None
    headers = [
        "" if str(h).startswith("Unnamed") else h
        for h in headers
    ]

    # --- Manual markdown rendering ---
    header_row = "| " + " | ".join(headers) + " |"
    separator_row = "| " + " | ".join(["---"] * len(headers)) + " |"

    body_rows = []
    for _, row in df.iterrows():
        body_rows.append(
            "| " + " | ".join(row.tolist()) + " |"
        )

    return "\n".join([header_row, separator_row] + body_rows)